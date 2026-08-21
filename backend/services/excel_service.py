import io
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def gerar_excel_validades(produtos: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()

    # --- ESTILOS COMPARTILHADOS ---
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")  # Azul Escuro
    font_header = Font(color="FFFFFF", bold=True, size=11)

    fill_coleta1 = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # Verde Claro
    fill_coleta2 = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")  # Azul Claro
    fill_coleta3 = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")  # Laranja Claro

    font_bold = Font(bold=True)
    font_divergencia = Font(bold=True, color="9C0006")
    fill_divergencia = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    thin_border = Side(style='thin', color='D9D9D9')
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    # ==========================================
    # ABA 1: DADOS DO ESTOQUE
    # ==========================================
    ws1 = wb.active
    ws1.title = "Dados do Estoque"

    headers_1 = [
        "Código", "EAN", "Complemento", "Descrição", "Separador",
        "Estoque_Sistema", "Reservado", "Disponível_Sistema"
    ]
    ws1.append(headers_1)

    for col_num in range(1, len(headers_1) + 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, p in enumerate(produtos, start=2):
        estoque_sistema = float(p.get("estoque", 0.0))
        reservado = float(p.get("reservado", 0.0))
        disponivel_sistema = float(p.get("disponivel", estoque_sistema - reservado))
        ean_val = str(p.get("ean", "")).strip()

        linha_1 = [
            p.get("cod", ""),
            ean_val,
            p.get("complemento", ""),
            p.get("desc", ""),
            p.get("separador", ""),
            estoque_sistema,
            reservado,
            disponivel_sistema
        ]
        ws1.append(linha_1)

        # Evita notação científica no EAN
        cell_ean = ws1.cell(row=row_idx, column=2)
        cell_ean.data_type = 's'
        cell_ean.number_format = '@'

        for col_num in range(1, len(headers_1) + 1):
            ws1.cell(row=row_idx, column=col_num).border = border

    # ==========================================
    # ABA 2: COLETA E LOTES
    # ==========================================
    ws2 = wb.create_sheet(title="Coleta e Lotes")

    headers_2 = [
        "Código", "Descrição", "Total_Coletado", "Divergência",
        "Qtd_1", "Data_1", "Qtd_2", "Data_2", "Qtd_3", "Data_3"
    ]
    ws2.append(headers_2)

    for col_num in range(1, len(headers_2) + 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, p in enumerate(produtos, start=2):
        coletas = p.get("coletas", [])

        qtd_1 = coletas[0].get("qtd", "") if len(coletas) > 0 else p.get("qtd_1", "")
        data_1 = coletas[0].get("data", "") if len(coletas) > 0 else p.get("data_1", "")

        qtd_2 = coletas[1].get("qtd", "") if len(coletas) > 1 else p.get("qtd_2", "")
        data_2 = coletas[1].get("data", "") if len(coletas) > 1 else p.get("data_2", "")

        qtd_3 = coletas[2].get("qtd", "") if len(coletas) > 2 else p.get("qtd_3", "")
        data_3 = coletas[2].get("data", "") if len(coletas) > 2 else p.get("data_3", "")

        estoque_sistema = float(p.get("estoque", 0.0))

        coletados_list = []
        for q in [qtd_1, qtd_2, qtd_3]:
            try:
                if q is not None and str(q).strip() != "":
                    coletados_list.append(float(str(q).replace(',', '.')))
            except ValueError:
                pass

        total_coletado = sum(coletados_list)
        divergencia = total_coletado - estoque_sistema

        linha_2 = [
            p.get("cod", ""),
            p.get("desc", ""),
            total_coletado,
            divergencia,
            qtd_1, data_1,
            qtd_2, data_2,
            qtd_3, data_3
        ]
        ws2.append(linha_2)

        # Destaque de cores nas coletas
        for col_i in [5, 6]:  # Qtd_1, Data_1
            c = ws2.cell(row=row_idx, column=col_i)
            c.fill = fill_coleta1
            c.font = font_bold

        for col_i in [7, 8]:  # Qtd_2, Data_2
            c = ws2.cell(row=row_idx, column=col_i)
            c.fill = fill_coleta2
            c.font = font_bold

        for col_i in [9, 10]:  # Qtd_3, Data_3
            c = ws2.cell(row=row_idx, column=col_i)
            c.fill = fill_coleta3
            c.font = font_bold

        # Destaque de divergência
        cell_div = ws2.cell(row=row_idx, column=4)
        if divergencia != 0:
            cell_div.fill = fill_divergencia
            cell_div.font = font_divergencia

        for col_num in range(1, len(headers_2) + 1):
            ws2.cell(row=row_idx, column=col_num).border = border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()